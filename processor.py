# processor.py
# Ablauf:
#   1) Stempelprotokoll-PDF parsen -> DataFrame ["Mitarbeiter","start","stop"] & Datum
#   2) Telematik (CSV/XLSX) lesen & pro Fahrer konsolidieren (früheste Startzeit, späteste Endzeit)
#   3) Join über normalisierten Namen, Abweichungen berechnen
#   4) Finale Excel mit bedingter Formatierung erzeugen
#
# WICHTIG: "start" = Wert unter Spaltenkopf "von", "stop" = Wert unter "bis" (aus der Tabelle im PDF)

import re
from io import BytesIO
from datetime import datetime
from typing import Optional, Tuple, List

import numpy as np
import pandas as pd
import pdfplumber
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.formatting.rule import CellIsRule, FormulaRule


# ----------------- Hilfsfunktionen -----------------
def norm_name(name: str) -> str:
    if name is None or (isinstance(name, float) and pd.isna(name)):
        return ""
    return re.sub(r"\s+", " ", str(name)).strip().lower()

def parse_hhmm(t):
    """Robuster HH:MM-Parser, akzeptiert z.B. '7:5' -> 07:05, sonst NaT."""
    if t is None or t == "" or (isinstance(t, float) and pd.isna(t)):
        return pd.NaT
    s = str(t).strip()
    m = re.match(r"^(\d{1,2}):(\d{1,2})$", s)
    if not m:
        return pd.NaT
    h = int(m.group(1)); mi = int(m.group(2))
    try:
        return datetime.strptime(f"{h:02d}:{mi:02d}", "%H:%M")
    except Exception:
        return pd.NaT

def diff_minutes(a,b):
    if pd.isna(a) or pd.isna(b):
        return np.nan
    return int((a-b).total_seconds()//60)

def hhmm_min_key(s: Optional[str]):
    dt = parse_hhmm(s)
    return (dt.hour if not pd.isna(dt) else 99, dt.minute if not pd.isna(dt) else 99)


# ----------------- PDF parsen (Name + von/bis -> start/stop) -----------------
def _find_name_and_left_date(page_text: str):
    """
    Kopfzeile z.B.:
    'Stempel-Protokoll Nachname, Vorname 29.09.2025 bis 28.10.2025'
    Liefert 'Vorname Nachname' und das linke Datum (z.B. 29.09.2025).
    """
    rx = re.compile(
        r"stempel[-\s]?protokoll\s+(.+?)\s+(\d{2}\.\d{2}\.\d{4})\s+bis\b",
        re.IGNORECASE,
    )
    m = rx.search(page_text or "")
    if not m:
        return None, None
    raw_name, left_date = m.groups()
    raw_name = raw_name.strip()
    if "," in raw_name:
        last, first = [x.strip() for x in raw_name.split(",", 1)]
        emp = f"{first} {last}"
    else:
        emp = raw_name
    return emp, left_date

def _get_lines(words, tol_y: float = 3.0):
    lines = []
    for w in sorted(words, key=lambda z: (z["top"], z["x0"])):
        for ln in lines:
            if abs(ln[0]["top"] - w["top"]) <= tol_y:
                ln.append(w); break
        else:
            lines.append([w])
    for ln in lines:
        ln.sort(key=lambda z: z["x0"])
    return lines

def _center_x(w): return (w["x0"] + w["x1"]) / 2.0

def _find_table_header_line(lines):
    """
    Finde die Zeile mit 'Datum' + 'von' + 'bis'.
    Rückgabe: (header_line, von_x, bis_x)
    """
    for ln in lines:
        tokens = [w["text"].strip().lower() for w in ln]
        if any("datum" in t for t in tokens) and ("von" in tokens) and ("bis" in tokens):
            von_x = None; bis_x = None
            for w in ln:
                t = w["text"].strip().lower()
                if t == "von" and von_x is None:
                    von_x = _center_x(w)
                elif t == "bis" and bis_x is None:
                    bis_x = _center_x(w)
            return ln, von_x, bis_x
    return None, None, None

def _pick_closest_time(line_words, target_x: float):
    cands = []
    for w in line_words:
        t = w.get("text","").strip()
        if re.match(r"^\d{1,2}:\d{2}$", t):
            cands.append((abs(_center_x(w) - target_x), t))
    if not cands:
        return None
    cands.sort(key=lambda x: x[0])
    return cands[0][1]

def _extract_start_stop_for_date(page, target_date: str):
    """
    1) Zeilen bilden, 2) Headerzeile (Datum/von/bis) finden → von_x/bis_x,
    3) alle Zeilen mit target_date suchen, je Zeile Zeit unter 'von'/'bis' picken,
    4) über mehrere Zeilen aggregieren: früheste 'start', späteste 'stop'.
    """
    words = page.extract_words(use_text_flow=True, keep_blank_chars=False) or []
    lines = _get_lines(words)

    header_ln, von_x, bis_x = _find_table_header_line(lines)
    if header_ln is None or von_x is None or bis_x is None:
        return None, None

    starts, stops = [], []
    for ln in lines:
        if not any(target_date in (w["text"] or "") for w in ln):
            continue
        s = _pick_closest_time(ln, von_x)
        e = _pick_closest_time(ln, bis_x)
        if s: starts.append(s)
        if e: stops.append(e)

    start = min(starts, key=hhmm_min_key) if starts else None
    stop  = max(stops,  key=hhmm_min_key) if stops  else None
    return start, stop

def parse_stempel_pdf_to_df(pdf_bytes: bytes, explicit_date: Optional[str] = None):
    """
    Liefert df_stempel: ["Mitarbeiter","start","stop"] und verwendetes Datum (TT.MM.JJJJ).
    """
    rows = []
    used_date = None

    with pdfplumber.open(BytesIO(pdf_bytes)) as pdf:
        for page in pdf.pages:
            text = page.extract_text() or ""
            emp, left_date = _find_name_and_left_date(text)
            if not emp or not left_date:
                continue

            date_str = explicit_date or left_date
            st, sp = _extract_start_stop_for_date(page, date_str)
            rows.append({"Mitarbeiter": emp, "start": st, "stop": sp})

            if used_date is None:
                used_date = date_str

    if not rows:
        raise ValueError("Keine 'von'/'bis'-Werte im PDF gefunden. Falls das PDF gescannt ist, wird eine Textschicht/OCR benötigt.")

    df = pd.DataFrame(rows)
    df["match_key"] = df["Mitarbeiter"].map(norm_name)
    return df, used_date or (explicit_date or "")


# ----------------- Telematik einlesen & konsolidieren -----------------
def read_telematik(file_bytes: bytes, filename: str) -> pd.DataFrame:
    """
    CSV oder Excel möglich. Erwartete Spalten:
      - Name Fahrer / Name
      - Startzeit
      - Endzeit
    Konsolidierung pro Fahrer: früheste Startzeit, späteste Endzeit.
    """
    fn = filename.lower()
    if fn.endswith(".csv"):
        text = file_bytes.decode("utf-8", errors="ignore")
        lines = [ln for ln in text.splitlines() if ln.strip()]
        if not lines:
            raise ValueError("Telematik-Datei ist leer.")
        header = [h.strip().lstrip("\ufeff") for h in lines[0].split(",")]
        rows = [ln.split(",") for ln in lines[1:]]
        df = pd.DataFrame(rows, columns=header)
    elif fn.endswith(".xlsx") or fn.endswith(".xlsm") or fn.endswith(".xls"):
        df = pd.read_excel(BytesIO(file_bytes))
    else:
        raise ValueError("Telematik-Datei muss CSV oder Excel sein.")

    df = df.rename(columns=lambda c: str(c).strip().lstrip("\ufeff"))

    name_col = next((c for c in ["Name Fahrer", "Name", "Fahrer", "Fahrer Name", "Fahrername"] if c in df.columns), None)
    start_col = next((c for c in ["Startzeit", "Start", "Zündung Start", "Ignition Start"] if c in df.columns), None)
    end_col   = next((c for c in ["Endzeit", "Ende", "Zündung Ende", "Ignition End"] if c in df.columns), None)
    if not name_col or not start_col or not end_col:
        raise ValueError("Telematik benötigt Spalten: Name + Startzeit + Endzeit.")

    tele = df[[name_col, start_col, end_col]].copy()
    tele.columns = ["Name", "Startzeit", "Endzeit"]

    def to_hhmm(v):
        if isinstance(v, str) and re.match(r"^\s*\d{1,2}:\d{2}\s*$", v):
            return v.strip()
        try:
            dt = pd.to_datetime(v, errors="coerce")
            if not pd.isna(dt):
                return dt.strftime("%H:%M")
        except Exception:
            pass
        return None

    tele["Startzeit"] = tele["Startzeit"].map(to_hhmm)
    tele["Endzeit"]   = tele["Endzeit"].map(to_hhmm)

    tele["match_key"] = tele["Name"].map(norm_name)

    grouped = tele.groupby("match_key").agg({
        "Name": "first",
        "Startzeit": lambda s: min([x for x in s if x], default=None, key=hhmm_min_key),
        "Endzeit":   lambda s: max([x for x in s if x], default=None, key=hhmm_min_key),
    }).reset_index()

    return grouped[["match_key", "Name", "Startzeit", "Endzeit"]]


# ----------------- Finale Excel (mit deinen Regeln) -----------------
def build_final_excel_bytes(out_df: pd.DataFrame, date_str: str) -> bytes:
    """
    Regeln:
      - Abweichung Stop: >10 rot; >5 und <10 gelb
      - Abweichung Start: < -10 rot; -10 < x < -5 gelb
      - start vor 07:30 rot
      - stop spät: Mo–Do > 17:15 rot; Fr > 13:15 rot
    """
    dt = datetime.strptime(date_str, "%d.%m.%Y")
    weekday = dt.weekday()  # 0=Mo ... 4=Fr

    wb = Workbook()
    ws = wb.active
    ws.title = "Auswertung"

    ws["A1"] = f"Datum: {date_str}"
    ws["A1"].font = Font(bold=True)
    ws["A1"].alignment = Alignment(horizontal="left")

    headers = [
        "Mitarbeiter",
        "start",
        "stop",
        "Startzeit Zündung",
        "Endzeit Zündung",
        "Abweichung Start [min]",
        "Abweichung Stop [min]",
    ]
    start_row = 3
    for j, h in enumerate(headers, start=1):
        ws.cell(row=start_row, column=j, value=h).font = Font(bold=True)

    df_to_write = out_df.rename(columns={"Startzeit": "Startzeit Zündung", "Endzeit": "Endzeit Zündung"})
    for r_idx, row in enumerate(df_to_write.itertuples(index=False), start=start_row + 1):
        for c_idx, val in enumerate(row, start=1):
            ws.cell(row=r_idx, column=c_idx, value=(None if pd.isna(val) else val))

    # Spaltenbreiten
    for i, w in enumerate([30, 12, 12, 16, 16, 22, 22], start=1):
        ws.column_dimensions[chr(64 + i)].width = w

    data_start = start_row + 1
    data_end   = start_row + len(df_to_write)

    red = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    yellow = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")

    # Abweichung Stop (Spalte G)
    ws.conditional_formatting.add(f"G{data_start}:G{data_end}", CellIsRule(operator="greaterThan", formula=["10"], fill=red))
    ws.conditional_formatting.add(f"G{data_start}:G{data_end}", FormulaRule(formula=[f"AND(G{data_start}>5,G{data_start}<10)"], fill=yellow))

    # Abweichung Start (Spalte F)
    ws.conditional_formatting.add(f"F{data_start}:F{data_end}", CellIsRule(operator="lessThan", formula=["-10"], fill=red))
    ws.conditional_formatting.add(f"F{data_start}:F{data_end}", FormulaRule(formula=[f"AND(F{data_start}>-10,F{data_start}<-5)"], fill=yellow))

    # start vor 07:30 (Spalte B)
    ws.conditional_formatting.add(f"B{data_start}:B{data_end}", FormulaRule(formula=[f"TIMEVALUE(B{data_start})<TIME(7,30,0)"], fill=red))

    # stop-Schwelle (Spalte C)
    stop_threshold = "TIME(13,15,0)" if weekday == 4 else "TIME(17,15,0)"
    ws.conditional_formatting.add(f"C{data_start}:C{data_end}", FormulaRule(formula=[f"TIMEVALUE(C{data_start})>{stop_threshold}"], fill=red))

    ws.freeze_panes = "A4"

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio.getvalue()


# ----------------- Öffentlicher Entry: NUR FINALE AUSWERTUNG -----------------
def generate_final_excel(
    pdf_bytes: bytes,
    tele_bytes: bytes,
    tele_filename: str,
    explicit_date: Optional[str] = None
):
    # 1) PDF -> df_stempel
    df_stempel, date_str = parse_stempel_pdf_to_df(pdf_bytes, explicit_date=explicit_date)

    # 2) Telematik lesen/konsolidieren
    df_tele = read_telematik(tele_bytes, tele_filename)

    # 3) Merge & Abweichungen
    merged = pd.merge(df_stempel, df_tele, on="match_key", how="outer", suffixes=("_stempel", "_tele"))
    merged["Mitarbeiter"] = merged["Mitarbeiter"].fillna(merged.get("Name"))
    if "Name" in merged.columns:
        merged = merged.drop(columns=["Name"])

    merged["t_start"] = merged["start"].map(parse_hhmm)
    merged["t_stop"]  = merged["stop"].map(parse_hhmm)
    merged["t_ign_s"] = merged["Startzeit"].map(parse_hhmm)
    merged["t_ign_e"] = merged["Endzeit"].map(parse_hhmm)

    merged["Abweichung Start [min]"] = merged.apply(lambda r: diff_minutes(r["t_start"], r["t_ign_s"]), axis=1)
    merged["Abweichung Stop [min]"]  = merged.apply(lambda r: diff_minutes(r["t_stop"],  r["t_ign_e"]), axis=1)

    out = merged[[
        "Mitarbeiter",
        "start",
        "stop",
        "Startzeit",
        "Endzeit",
        "Abweichung Start [min]",
        "Abweichung Stop [min]",
    ]].copy().sort_values("Mitarbeiter").reset_index(drop=True)

    # 4) Finale Excel
    final_xlsx = build_final_excel_bytes(out, date_str)
    final_name = f"Auswertung_Stempel_vs_Telematik_{date_str}.xlsx"
    return final_xlsx, final_name
